import sys
from pathlib import Path

import openpyxl


def find_header_row(ws, min_non_empty: int = 3, max_scan_rows: int = 50):
    for r in range(1, min(ws.max_row, max_scan_rows) + 1):
        values = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        cleaned = [str(v).strip() for v in values if v is not None and str(v).strip() != ""]
        if len(cleaned) >= min_non_empty:
            # Stop at first likely header row.
            return r
    return None


def main(argv: list[str]) -> int:
    if len(argv) < 2:
        print("Usage: python tools/extract_cfu_schema_from_xlsx.py <path-to-xlsx>")
        return 2

    xlsx_path = Path(argv[1]).expanduser().resolve()
    if not xlsx_path.exists():
        print(f"ERROR: not found: {xlsx_path}")
        return 2

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    sheet = None
    for name in wb.sheetnames:
        if str(name).strip().lower() == "cfu":
            sheet = wb[name]
            break

    if sheet is None:
        print("ERROR: no sheet named 'cfu' found.")
        print("Sheets:")
        for n in wb.sheetnames:
            print("-", n)
        return 1

    header_row = find_header_row(sheet)
    if header_row is None:
        print("ERROR: could not find a header row.")
        return 1

    headers = []
    for c in range(1, sheet.max_column + 1):
        v = sheet.cell(row=header_row, column=c).value
        s = "" if v is None else str(v).strip()
        if s == "":
            # stop after first long run of empties (common in templates)
            # but keep a couple empties in case of gaps
            pass
        headers.append(s)

    # Trim trailing empties
    while headers and headers[-1] == "":
        headers.pop()

    print(f"Workbook: {xlsx_path}")
    print(f"Sheet: {sheet.title}")
    print(f"Header row: {header_row}")
    print("\nColumns (one per line):")
    for h in headers:
        print(h)

    print("\nJS array literal:")
    js = ",\n  ".join([repr(h) for h in headers])
    print("[\n  " + js + "\n]")

    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
