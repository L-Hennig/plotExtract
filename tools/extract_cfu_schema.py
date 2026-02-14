import argparse
from pathlib import Path

import openpyxl


def _as_str(v):
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    return str(v).strip()


def extract_headers(xlsx_path: Path, sheet_name: str) -> list[str]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise SystemExit(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
    ws = wb[sheet_name]

    # Heuristic: header row is first row with >= 3 non-empty cells.
    for row_idx in range(1, min(50, ws.max_row or 1) + 1):
        row_vals = [_as_str(ws.cell(row=row_idx, column=c).value) for c in range(1, (ws.max_column or 1) + 1)]
        # Trim trailing empties
        while row_vals and row_vals[-1] == "":
            row_vals.pop()
        non_empty = [v for v in row_vals if v]
        if len(non_empty) >= 3:
            return row_vals

    return []


def main() -> None:
    ap = argparse.ArgumentParser(description="Print the header columns from an Excel sheet (default: 'cfu').")
    ap.add_argument("xlsx", type=Path, help="Path to .xlsx template")
    ap.add_argument("--sheet", default="cfu", help="Sheet name (default: cfu)")
    args = ap.parse_args()

    xlsx_path: Path = args.xlsx
    if not xlsx_path.exists():
        raise SystemExit(f"File not found: {xlsx_path}")

    headers = extract_headers(xlsx_path, args.sheet)
    if not headers:
        raise SystemExit("No header row found (looked in first 50 rows).")

    print(f"TEMPLATE: {xlsx_path}")
    print(f"SHEET: {args.sheet}")
    print(f"COLUMNS ({len(headers)}):")
    for i, h in enumerate(headers, start=1):
        print(f"{i:02d}. {h}")


if __name__ == "__main__":
    main()
